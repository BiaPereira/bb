import os
import numpy as np
import matplotlib.pyplot as plt
from bisect import bisect
from lmfit.models import GaussianModel, ConstantModel
from scipy.odr import *
from lmfit import Model
import openpyxl
from math import sin, acos, cos, fabs

greek_alphabet = {u'\u03A7'}

#Opens file
filename = input('Name of the file: ')
#filename = 'F2LiCl_001.txt'
tam = len(filename)

if filename[tam-3:tam] == 'txt':
    file = open(filename)
if filename[tam-3:tam] == 'odf':
    base = os.path.splitext(filename)[0]
    os.rename(filename, base + ".txt")
    file = open (filename[0:tam-3]+'txt')
fileR = file.read()

#def

def grafico (x,y):
    # legend
    plt.plot(x, y, 'b+',label="Experimental")
    plt.legend(bbox_to_anchor=(0.8, 1), loc=2, borderaxespad=0.)
    plt.yscale('log', nonposy='clip')
    plt.ylabel('Intensidade')
    plt.xlabel('Canal')
    plt.show()


#function
def zoom_gráfico (minimo, maximo):
    lmin = x.index(minimo)
    print('1')
    lmax = x.index(maximo)
    new_x = []
    new_y = []
    print('2')
    new_x = x[lmin:lmax + 1]
    new_y = y[lmin:lmax + 1]
    print('m')
    print (new_x)
    print(new_y)
    grafico(new_x, new_y)

#Gauss + line + bimodal

def k (M2,teta):
    M1 = 1.007825
    kli = ((((M2**2-M1**2*(sin(teta))**2)**(1/2)+(M1*cos(teta)))/(M1+M2))**2)
    return kli

def gaussian (x, A,index_x, FWHM):
    gauss = (A/((FWHM/(np.log(4))**(1/2))))*np.exp(-2*((x-index_x)**2)/((FWHM/(np.log(4))**(1/2))**2))
    return gauss

def line (x, m, b):
    return m*x+b

def bimodal(x,A1,index1,FWHM1,A2,index2,FWHM2,teta, k6, k7):
    FWHM1 = fabs(FWHM1)
    FWHM2 = fabs(FWHM2)
    return gaussian(x,A1,index1,FWHM1)+ gaussian(x,A2,index2,FWHM2)

def graf(xtotal, ytotal, mod, pars, file):
    result = mod.fit(ytotal, pars, x=xtotal, weights=1 / np.sqrt(ytotal))
    print(result.fit_report())
    report = result.fit_report().split('\n', -1)
    report7 = report[7]
    lreport = len(report7)
    valor = report7[24:lreport]
    plt.plot(xtotal, ytotal, 'b+', label="Experimental")
    plt.plot(xtotal, result.init_fit, 'k--', color='grey', label="Initial Fit")
    plt.plot(xtotal, result.best_fit, 'r-', color="green", label="Best Fit")
    plt.yscale('log', nonposy='clip')
    plt.ylabel('Intensidade')
    plt.xlabel('Canal')
    plt.figtext(.139, 0.800, r'$\frac{\chi^{2}}{NDF}$', fontsize=13)
    plt.figtext(.19, 0.8, '=' + valor)
    plt.legend(bbox_to_anchor=(0.8, 1), loc=2, borderaxespad=0.)
    plt.show()

    newfile = input('file name:')

    if os.path.exists(newfile):
        y = open(newfile, 'a')
        y.write('\n')
        y.write (file)
        y.write('\n')
        y.write("\n".join(map(lambda x: str(x), report[11:17])))
        y.write(' ')
        y.close()
    else:
        f = open(newfile, 'w')
        f.write('\n')
        f.write (file)
        f.write('\n')
        f.write("\n".join(map(lambda x: str(x), report[11:17])))
        f.write(' ')
        f.close()
#print(resul.fit_report())
#t=resul.fit_report().split('\n',-1)
#print (t)
#l = t[7]
#tamanho = len(l)
#valorchi = l[24:tamanho]

#split
g = fileR.split(' ',-1)
leng = len(g)

#Two list
i=0
x = []
y = []
print(type(g))
for i in range(0,leng-1,4):

    x.append(float(g[i + 1]))
    y.append(float(g[i+3]))

#Max and min vector x and y
max_x = max(x)
min_x = min(x)
max_y = max(y)
min_y = min(y)

# Plot
plt.grid()
grafico(x,y)

#---------------------New plot
# input limit plot
lim_min = int(input('Lower limit: '))
lim_max = int(input('Upper limit: ' ))


zoom_gráfico(lim_min,lim_max)
resp = input('Gráfico Correto? Sim/Não: ')
while resp == 'Não':
    lim_min = int(input('Lower limit: '))
    lim_max = int(input('Upper limit: '))
    zoom_gráfico(lim_min,lim_max)
    resp = input('Gráfico Correto? Sim/Não: ')

print ('Entre que valores está o pico2 máximo?')
min_xx = int(input('Lower limit: '))
max_xx = int(input('Upper limit: '))
print ('segundo pico')
min_xx2 = int(input('Lower limit: '))
max_xx2 = int(input('Upper limit: '))
#y0 = int(input('y0:'))
#min_xx = 2530
#max_xx = 2640
#min_xx2 = 2630
#max_xx2 = 2740

if min_xx<min_xx2:
    xtotal = x[min_xx:max_xx2 + 1]
    ytotal= y[min_xx:max_xx2 + 1]
if min_xx>min_xx2:
    xtotal = x[min_xx2:max_xx + 1]
    ytotal = y[min_xx2:max_xx + 1]

# New vector x and y (shorter)
xx = x[min_xx:max_xx + 1]
yy = y[min_xx :max_xx + 1]

xx2 = x[min_xx2:max_xx2 + 1]
yy2 = y[min_xx2:max_xx2 + 1]

grafico(xtotal,ytotal)

print('yy', yy)
print(type(yy))

#Obter os valores
max_yy = max(yy)#max_counts
FWHMyy = max_yy/2 #number of FWHM yy
#print ('FWHMyy',FWHMyy)
index_max_yy = yy.index(max_yy)
index_max_y = index_max_yy + min_xx
#In case that FWHM does not appear in yy
print(index_max_yy)
yye=yy[index_max_yy-30:index_max_yy+30]
print('yy3',yye)
print(FWHMyy)
if FWHMyy in yye:
    index_FWHM_y = yye.index(FWHMyy) + min_xx + index_max_yy-30

else:
    y2 = sorted(yye) #order
    print (y2)
    index_FWHM_y2 = bisect(y2,FWHMyy)
    #print ('index_FWHM_y2',index_FWHM_y2)
    number = y2[index_FWHM_y2]
    #print (number)
    index_FWHM_y = yye.index(number) + min_xx + index_max_yy-30
print(index_FWHM_y)

max_yy2 = max(yy2)#max_counts

FWHMyy2 = max_yy2/2 #number of FWHM yy
index_max_y2 = yy2.index(max_yy2) + min_xx2
print(FWHMyy2)
if FWHMyy2 in yy2:
    index_FWHM_y2 = yy2.index(FWHMyy2) + min_xx2

else:
    y2 = sorted(yy2) #order
    print (y2)
    index_FWHM_y2 = bisect(y2,FWHMyy2)
    print ('index_FWHM_y2',index_FWHM_y2)
    number = y2[index_FWHM_y2]
    print (number)
    index_FWHM_y2 = yy2.index(number) + min_xx2
    print(yy2.index(number))




#print ('yy.index(maxy)',yy.index(max_yy))
#print ('yy.index(FWHMy)', index_FWHM_y)


FWHM_xx = (abs(index_max_y-index_FWHM_y)*2)
print(FWHM_xx)
#print ('FWHM', FWHM_xx)
FWHM2 = (abs(index_max_y-index_FWHM_y))
print(FWHM2)
FWHM_xx2 = (abs(index_max_y2-index_FWHM_y2)*2)
print(FWHM_xx2)
FWHM22 = (abs(index_max_y2-index_FWHM_y2))
print(FWHM22)

#Area
inicio = index_max_y - min_xx - FWHM2
fim = inicio + FWHM2
#print(yy)
# print (inicio)
#print (fim)
area = 0
j = inicio
for j in range(inicio, fim,1):
    area = (area + yy[j])
    #print ('area1:', area)

inicio2 = index_max_y2 - min_xx2 - FWHM22
fim2 = inicio2 + FWHM_xx2
area2 = 0
g = inicio2
for g in range(inicio2, fim2,1):
    area2 = (area2 + yy2[g])

xx=np.array(xx, dtype=float)
yy=np.array(yy, dtype=float)
xx2=np.array(xx2, dtype=float)
yy2=np.array(yy2, dtype=float)
xtotal=np.array(xtotal, dtype=float)
ytotal=np.array(ytotal, dtype=float)

mod = Model(bimodal) + Model(line)
#plt.plot(bimodal(x=xtotal, A1= area, index1=index_max_y, FWHM1=FWHMyy,A2 = area2, index2= index_max_y2, FWHM2= FWHMyy2))
pars = mod.make_params(A1= area, index1=index_max_y, FWHM1=FWHM_xx,A2 = area2, index2= index_max_y2, FWHM2= FWHM_xx2,
                       teta=0, k6=0, k7=0, m=0, b=10)
pars['FWHM1'].min = 0.001         # sigma  > 0
pars['FWHM2'].min = 0.001     # amplitude > 0
pars['k6'].set(vary=False)
pars['k7'].set(vary=False)
pars['teta'].set(vary=False)

graf(xtotal, ytotal, mod, pars, filename)

#book = openpyxl.load_workbook(input('Ficheiro Excel'))
book = openpyxl.load_workbook('LiCl_Logbook.xlsx')

sheet = book.active

# Valor do angulo
for i in range(1,100,1):
    loc = sheet.cell(row=i, column=25)
    if loc.value==filename[2:tam-4]:
        print (loc)
        lin = loc.row
        if filename[0:2] == 'F2':
            teta = sheet.cell(row= lin, column = 25-2)
        if filename[0:2] == 'F3':
            teta = sheet.cell(row=lin, column=25 - 1)
teta = teta.value
print(teta)


m6 = 6.015121 #"* ma
m7 = 7.016005 #* ma


#Calculo dos k
k6 = k(m6,teta)
print(k6)
k7 = k(m7,teta)
print(k7)

mod = Model(bimodal) + Model(line)

pars = mod.make_params(A1=area, index1=index_max_y, FWHM1=FWHM_xx, A2=area2, index2=index_max_y2, FWHM2=FWHM_xx2,
                       teta=teta, k6=k6, k7=k7, m=-5, b=10)         # sigma  > 0
pars['k6'].set(vary=False)
pars['k7'].set(vary=False)
pars['FWHM1'].set(expr='(((k6 + (1/cos(teta)))/(k7 + (1/cos(teta))))*FWHM2)')
pars['teta'].set(vary=False)


graf(xtotal, ytotal, mod, pars, filename)

anw= input('Guardar? s/n')
if anw == 's':
    limites = 'limites'
    lim = open(limites, 'a')
    lim.write('\n')
    lim.write(filename)
    lim.write(': ')
    lim.write(str(min_xx))
    lim.write(' ')
    lim.write(str(max_xx))
    lim.write(' ')
    lim.write(str(min_xx2))
    lim.write(' ')
    lim.write(str(max_xx2))
    lim.write(' ')
    lim.write('\n')
    lim.close()

